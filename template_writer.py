from __future__ import annotations
import os
import re
import shutil
from copy import copy
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class TemplateWriter:
    """
    DataFrame → 模板写回（保样式），与原脚本一致；支持覆盖前自动备份。
    """

    @staticmethod
    def _normalize(name: str) -> str:
        if name is None:
            return ""
        s = "".join(ch for ch in str(name).strip() if ch.isalnum())
        return s.lower()

    @staticmethod
    def _read_header_row(ws: Worksheet, row_idx: int) -> List[str]:
        max_col = ws.max_column
        return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]

    def _build_mapping_from_df(self, df_columns: List[str], tpl_headers: List[str]) -> Dict[int, int]:
        src_map = {self._normalize(h): i + 1 for i,
                   h in enumerate(df_columns) if h}
        mapping: Dict[int, int] = {}
        for j, h in enumerate(tpl_headers, start=1):
            key = self._normalize(h)
            if key and key in src_map:
                mapping[src_map[key]] = j
        return mapping

    @staticmethod
    def _copy_style(dst_cell, src_cell):
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)

    @staticmethod
    def _next_backup_name(base: str) -> str:
        if not os.path.exists(base):
            return base
        root, ext = os.path.splitext(base)
        n = 1
        while True:
            cand = f"{root}({n}){ext}"
            if not os.path.exists(cand):
                return cand
            n += 1

    @staticmethod
    def _sanitize_excel(s: Any) -> Any:
        if s is None:
            return ""
        try:
            if pd.isna(s):
                return ""
        except Exception:
            pass
        s = str(s)
        if s.strip().lower() == "nan":
            return ""
        return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", s)

    def fill_template_with_df(
        self,
        df: pd.DataFrame,
        template_path: str,
        save_path: Optional[str] = None,
        do_backup_if_overwrite: bool = True,
        sheet_name: Optional[str] = None,
    ) -> Tuple[int, int, Optional[str]]:
        tpl_wb = load_workbook(template_path)
        if sheet_name and sheet_name in tpl_wb.sheetnames:
            tpl_ws = tpl_wb[sheet_name]
        else:
            tpl_ws = tpl_wb[tpl_wb.sheetnames[0]]

        tpl_headers_en = self._read_header_row(tpl_ws, 2)
        mapping = self._build_mapping_from_df(list(df.columns), tpl_headers_en)
        if not mapping:
            raise RuntimeError("未在模板第2行英文表头中找到与 DataFrame 列名对应的映射。")

        evidencet_col_idx = None
        for j, hdr in enumerate(tpl_headers_en, start=1):
            if self._normalize(str(hdr)) == "evidencet":
                evidencet_col_idx = j
                break

        backup_path: Optional[str] = None
        if save_path is None and do_backup_if_overwrite:
            root, _ext = os.path.splitext(template_path)
            backup_path_raw = f"{root}_备份.xlsx"
            backup_path = self._next_backup_name(backup_path_raw)
            shutil.copy2(template_path, backup_path)

        style_row_idx = 3
        write_count = 0
        for r_idx in range(len(df)):
            r_dst = style_row_idx + write_count
            for src_col_idx, tpl_col_idx in mapping.items():
                val_raw = df.iloc[r_idx, src_col_idx - 1]
                val = self._sanitize_excel(val_raw)
                dst_cell = tpl_ws.cell(row=r_dst, column=tpl_col_idx)
                dst_cell.value = val
                style_cell = tpl_ws.cell(row=style_row_idx, column=tpl_col_idx)
                self._copy_style(dst_cell, style_cell)
            write_count += 1

        if evidencet_col_idx is not None:
            last_row = max(tpl_ws.max_row, style_row_idx + len(df) - 1)
            for r in range(style_row_idx, last_row + 1):
                tpl_ws.cell(row=r, column=evidencet_col_idx).value = ""

        target_path = save_path if save_path else template_path
        tpl_wb.save(target_path)
        return write_count, len(mapping), backup_path
